#!/usr/bin/env python3
"""
OS店铺数据大盘分析工具 (v3 - 列名匹配)
分析Shopee/Tokopedia店铺的库存偏差、发货时间异常、商品重量/名称异常。
不再依赖固定列号，改为通过表头名称匹配列位置。
"""
import sys
import io
import re
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict

# Force UTF-8 output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ── 配置 ──────────────────────────────────────────────
SHEET_NAME = 'Vkiau店铺折扣 库存 预售汇总6.24'

# 列名常量（用于查找和输出）
COL_PRODUCT_ID = "Product ID"
COL_PRODUCT_NAME = "Product Name(Optional)"
COL_VARIATION_ID = "Variation ID"
COL_VARIATION_NAME = "Variation name(Optional)"
COL_SKU_REF = "SKU Ref. No.(Optional)"
COL_SHIPPING_TIME = "Shipping time 发货时间"
COL_GAP200_INC = "判断是否增加库存Gap200"
COL_GAP200_DEC = "判断是否减少库存Gap200"
COL_GAP1000_INC = "判断是否增加库存Gap1000"
COL_IDR001_AVAIL = "店铺（可用量-待审订单预占）IDR001"
COL_IDR001_PLATFORM = "Inventory on the platform（IDR001）店铺后台库存"
COL_SBY001_AVAIL = "店铺（可用量-待审订单预占）SBY001"
COL_SBY001_PLATFORM = "Inventory on the platform（SBY001）店铺后台库存"
COL_WEIGHT = "重量"
COL_LENGTH = "长"
COL_WIDTH = "宽"
COL_HEIGHT = "高"
COL_SELL_DAYS = "Available selling days 可售天数"

BASE_COL_NAMES = [
    COL_PRODUCT_ID, COL_PRODUCT_NAME, COL_VARIATION_ID,
    COL_VARIATION_NAME, COL_SKU_REF, COL_SHIPPING_TIME,
]
BASE_HEADERS = [
    "Product ID", "Product Name(Optional)", "Variation ID",
    "Variation name(Optional)", "SKU Ref. No.(Optional)",
    "Shipping time 发货时间",
]

TAIL_COL_NAMES = [COL_WEIGHT, COL_LENGTH, COL_WIDTH, COL_HEIGHT]
TAIL_HEADERS = ["重量", "长", "宽", "高"]

IDR001_COL_NAMES = [COL_IDR001_AVAIL, COL_IDR001_PLATFORM]
IDR001_HEADERS = [
    "店铺（可用量-待审订单预占）IDR001",
    "Inventory on the platform（IDR001）店铺后台库存",
]

SBY001_COL_NAMES = [COL_SBY001_AVAIL, COL_SBY001_PLATFORM]
SBY001_HEADERS = [
    "店铺（可用量-待审订单预占）SBY001",
    "Inventory on the platform（SBY001）店铺后台库存",
]


# ── 列名映射 ──────────────────────────────────────────
def build_column_map(ws):
    """读取第一行表头，建立 列名→索引(0-based) 的映射（自动规范化空白字符）"""
    col_map = {}
    for i, cell in enumerate(ws[1]):
        if cell.value:
            name = re.sub(r'\s+', ' ', str(cell.value).strip())
            col_map[name] = i
    return col_map


# ── 工具函数 ──────────────────────────────────────────
def safe_float(v):
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def is_true(v):
    return str(v).upper() == 'TRUE'


def extract_rows(ws, filter_func, col_map, out_col_names):
    """遍历工作表，筛选符合条件的行，返回输出列的值列表。"""
    out_indices = [col_map[n] for n in out_col_names]
    results = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if filter_func(row, col_map):
            results.append(tuple(row[i] for i in out_indices))
    return results


# ── 过滤条件 ──────────────────────────────────────────
def gap_increase_200(r, col_map):
    """店铺库存比ERP少200以上"""
    return is_true(r[col_map[COL_GAP200_INC]])


def gap_decrease_200(r, col_map):
    """店铺库存比ERP多200以上"""
    return is_true(r[col_map[COL_GAP200_DEC]])


def gap_increase_1000(r, col_map):
    """店铺库存比ERP少1000以上"""
    return is_true(r[col_map[COL_GAP1000_INC]])


def shipping_anomaly(r, col_map):
    """可用量>300 但发货时间>2"""
    ad = safe_float(r[col_map[COL_IDR001_AVAIL]])
    st = safe_float(r[col_map[COL_SHIPPING_TIME]])
    if ad is None or st is None:
        return False
    return ad > 300 and st > 2


def weight_anomaly(r, col_map):
    """可用量>300 且重量>10000"""
    ad = safe_float(r[col_map[COL_IDR001_AVAIL]])
    wt = safe_float(r[col_map[COL_WEIGHT]])
    if ad is None or wt is None:
        return False
    return ad > 300 and wt > 10000


def habis_anomaly(r, col_map):
    """可用量>300 且Variation name含habis"""
    ad = safe_float(r[col_map[COL_IDR001_AVAIL]])
    if ad is None:
        return False
    vname = str(r[col_map[COL_VARIATION_NAME]]) if r[col_map[COL_VARIATION_NAME]] is not None else ''
    return ad > 300 and 'habis' in vname.lower()


def sby001_adjust(r, col_map):
    """SBY001可用量>500"""
    af = safe_float(r[col_map[COL_SBY001_AVAIL]])
    return af is not None and af > 500


# 子表配置: (子表名, 过滤函数, 输出列名列表, 表头列表)
SHEETS_CONFIG = [
    ("增加库存Gap200", gap_increase_200, BASE_COL_NAMES + IDR001_COL_NAMES + TAIL_COL_NAMES, BASE_HEADERS + IDR001_HEADERS + TAIL_HEADERS),
    ("减少库存Gap200", gap_decrease_200, BASE_COL_NAMES + IDR001_COL_NAMES + TAIL_COL_NAMES, BASE_HEADERS + IDR001_HEADERS + TAIL_HEADERS),
    ("增加库存Gap1000", gap_increase_1000, BASE_COL_NAMES + IDR001_COL_NAMES + TAIL_COL_NAMES, BASE_HEADERS + IDR001_HEADERS + TAIL_HEADERS),
    ("发货时间异常", shipping_anomaly, BASE_COL_NAMES + IDR001_COL_NAMES + TAIL_COL_NAMES, BASE_HEADERS + IDR001_HEADERS + TAIL_HEADERS),
    ("重量异常_大于10000", weight_anomaly, BASE_COL_NAMES + IDR001_COL_NAMES + TAIL_COL_NAMES, BASE_HEADERS + IDR001_HEADERS + TAIL_HEADERS),
    ("名称含Habis异常", habis_anomaly, BASE_COL_NAMES + IDR001_COL_NAMES + TAIL_COL_NAMES, BASE_HEADERS + IDR001_HEADERS + TAIL_HEADERS),
    ("SBY001库存调整", sby001_adjust, BASE_COL_NAMES + SBY001_COL_NAMES + TAIL_COL_NAMES, BASE_HEADERS + SBY001_HEADERS + TAIL_HEADERS),
]


# ── 广告操作判断（计算型）────────────────────────────────
def compute_ad_judgment(ws, col_map):
    """按PID分组，统计MID数量和可售天数<=7的MID数量"""
    data = defaultdict(lambda: {'mids': set(), 'lte7': set()})
    idx_pid = col_map[COL_PRODUCT_ID]
    idx_mid = col_map[COL_VARIATION_ID]
    idx_sell_days = col_map[COL_SELL_DAYS]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        pid = row[idx_pid]
        mid = row[idx_mid]
        sell_days = safe_float(row[idx_sell_days])

        if pid is None:
            continue
        pid = str(pid)
        if mid is not None:
            data[pid]['mids'].add(mid)
            if sell_days is not None and sell_days <= 7:
                data[pid]['lte7'].add(mid)

    results = [(pid, len(d['mids']), len(d['lte7']))
               for pid, d in sorted(data.items())]
    return results


def compute_mid_sell_days(ws, col_map):
    """列出每个PID下每个MID及其可售天数"""
    idx_pid = col_map[COL_PRODUCT_ID]
    idx_mid = col_map[COL_VARIATION_ID]
    idx_sell_days = col_map[COL_SELL_DAYS]

    results = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        pid = row[idx_pid]
        mid = row[idx_mid]
        sell_days = safe_float(row[idx_sell_days])

        if pid is not None and mid is not None:
            results.append((str(pid), str(mid), sell_days if sell_days is not None else ''))

    results.sort(key=lambda x: (x[0], x[1]))
    return results


COMPUTED_SHEETS = [
    ("广告操作判断", compute_ad_judgment,
     ["PID", "MID数量", "可售天数小于等于7的MID数量"]),
    ("MID可售天数明细", compute_mid_sell_days,
     ["PID", "MID", "MID可售天数"]),
]


# ── 输出 ──────────────────────────────────────────────
def write_sheet(ws_out, data_rows, headers):
    """写入一个子表的数据。"""
    hfont = Font(bold=True, size=11, color='FFFFFF')
    hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    for ci, h in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=ci, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for ri, row_data in enumerate(data_rows, 2):
        for ci, val in enumerate(row_data, 1):
            ws_out.cell(row=ri, column=ci, value=val)

    for ci in range(1, len(headers) + 1):
        ml = max(len(str(headers[ci - 1])), 12)
        for ri in range(2, min(len(data_rows) + 2, 50)):
            cv = ws_out.cell(row=ri, column=ci).value
            if cv:
                ml = max(ml, min(len(str(cv)), 40))
        ws_out.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = ml + 3


def build_workbook(input_path):
    """读取输入文件，执行所有分析，返回输出工作簿。"""
    wb = openpyxl.load_workbook(input_path, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        available = '\n  - '.join(wb.sheetnames)
        raise ValueError(
            f'找不到工作表 "{SHEET_NAME}"。\n'
            f'可用的工作表:\n  - {available}'
        )

    ws = wb[SHEET_NAME]
    col_map = build_column_map(ws)
    print(f'  📋 找到 {len(col_map)} 个列')

    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    total = 0
    for sheet_name, filter_func, out_col_names, headers in SHEETS_CONFIG:
        rows = extract_rows(ws, filter_func, col_map, out_col_names)
        total += len(rows)
        ws_out = out_wb.create_sheet(title=sheet_name)
        write_sheet(ws_out, rows, headers)
        print(f'  [{sheet_name}] {len(rows)} 行')

    for sheet_name, compute_func, headers in COMPUTED_SHEETS:
        rows = compute_func(ws, col_map)
        total += len(rows)
        ws_out = out_wb.create_sheet(title=sheet_name)
        write_sheet(ws_out, rows, headers)
        print(f'  [{sheet_name}] {len(rows)} 行')

    return out_wb, total


# ── 主入口 ────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description='OS店铺数据大盘分析工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            '示例:\n'
            '  %(prog)s                                          # 默认文件\n'
            '  %(prog)s --input "OS店铺基本数据大盘.xlsx"          # 指定输入\n'
            '  %(prog)s -i "数据.xlsx" -o "结果.xlsx"             # 指定输入输出\n'
            '  %(prog)s --output "结果.xlsx"                      # 仅指定输出\n'
        ),
    )
    parser.add_argument('--input', '-i', default='OS店铺基本数据大盘.xlsx',
                        help='输入Excel文件路径 (默认: OS店铺基本数据大盘.xlsx)')
    parser.add_argument('--output', '-o', default='OS店铺数据大盘_分析结果.xlsx',
                        help='输出Excel文件路径 (默认: OS店铺数据大盘_分析结果.xlsx)')
    args = parser.parse_args()

    print(f'🔍 读取: {args.input}')
    print(f'📊 分析工作表: {SHEET_NAME}')

    try:
        out_wb, total = build_workbook(args.input)
    except Exception as e:
        print(f'❌ 错误: {e}', file=sys.stderr)
        sys.exit(1)

    out_wb.save(args.output)
    print(f'\n✅ 完成！共 {total} 行结果')
    print(f'📁 输出: {args.output}')
    print(f'📑 子表 ({len(out_wb.sheetnames)}个): {", ".join(out_wb.sheetnames)}')


if __name__ == '__main__':
    main()
